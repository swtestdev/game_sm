import time
import tkinter as tk
from tkinter import messagebox
import socket
import win32com.client
import openpyxl

import threading
from time import sleep
from tkinter import ttk
import _thread
from tkinter import filedialog
import json
import os
import codecs
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

server = None
# game_id: int = 0
game_data = None
file_results = "Table_Ottawa.xlsx"
game_score = {}
# frame_games = {}
# 0/RED=Wrong, 1/GREEN=Good and 2/Yellow=None
# bg_result = ["#F61F03", "#2FFFC1", "#F6FCC7"]
# HOST_ADDR = socket.gethostbyname(socket.gethostname())  # 0.0.0.0
HOST_ADDR = socket.gethostname()
print(f"Host IP: {socket.gethostname()}")
print(f"Host IP: {socket.gethostbyname(socket.gethostname())}")
HOST_PORT = 8080
print(f"Port used: {HOST_PORT}")

client_name = " "
clients = []
clients_names = []
player_data = []


def open_socket():
    global server, HOST_ADDR, HOST_PORT  # , teams  # code is fine without this

    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    print(socket.AF_INET)
    print(socket.SOCK_STREAM)
    server.bind((HOST_ADDR, HOST_PORT))
    server.listen(10)  # server is listening for client connection


def list_to_str(data: list):
    my_str = ''
    if len(data) > 0:
        for i in data:
            if my_str == '' or i == '/' or i == '=' or i == '+' or my_str[-1] == '/':
                my_str += i
            else:
                my_str += f" {i}"
    return my_str


window = tk.Tk()
window.title("Game Server")

# open server socket
open_socket()
text = f"HOST: {socket.gethostname()}/{socket.gethostbyname(socket.gethostname())}\nPort: {HOST_PORT}"
btnFrame = tk.Frame(window)
btnStart = tk.Button(btnFrame, text="Start Server", command=lambda: start_server())
btnStart.pack(side=tk.LEFT)
btnGameLaoding = tk.Button(btnFrame, text="Open PPTX game file", font=("Arial Bold", 10), bg="orange", fg="red",
                           command=lambda: load_places(), state=tk.DISABLED)
btnGameLaoding.pack(side=tk.LEFT)
btnStop = tk.Button(btnFrame, text="Close Server", command=lambda: close_server(), state=tk.NORMAL)
btnStop.pack(side=tk.LEFT)
btnFrame.pack(side=tk.TOP, pady=(5, 0))

# Middle frame consisting of two labels for displaying the host and port info
middleFrame = tk.Frame(window)
lblHost = tk.Label(middleFrame, text=f"Address: {HOST_ADDR}")
lblHost.pack(side=tk.LEFT)
lblPort = tk.Label(middleFrame, text=f"Port: {HOST_PORT}")
lblPort.pack(side=tk.LEFT)
lblFile = tk.Label(middleFrame, text=f"File: {file_results}")
lblFile.pack(side=tk.LEFT)
middleFrame.pack(side=tk.TOP, pady=(5, 0))

# Game table
gameFrame = tk.Frame(window)
gameFrame.pack(side=tk.TOP, pady=(5, 0))

# The client frame shows the client area
clientFrame = tk.Frame(window)
lblLine = tk.Label(clientFrame, text="********** LOG **********").pack()
scrollBar = tk.Scrollbar(clientFrame)
scrollBar.pack(side=tk.RIGHT, fill=tk.Y)
tkDisplay = tk.Text(clientFrame, height=10, width=100)
tkDisplay.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0))
scrollBar.config(command=tkDisplay.yview)
tkDisplay.config(
    yscrollcommand=scrollBar.set,
    background="#F4F6F7",
    highlightbackground="grey",
    state="disabled",
)
clientFrame.pack(side=tk.BOTTOM, pady=(5, 10))
if text != '' or text is not None:
    tkDisplay.config(state=tk.NORMAL)
    tkDisplay.insert(tk.END, text + "\n")
tkDisplay.config(state=tk.DISABLED)


def start_server():
    if len(clients_names) == 0:
        # accept connection a first client required
        _thread.start_new_thread(accept_client_connection, (server,))
    else:
        message_get_excel_file()
        pass


def accept_client_connection(the_server):
    global game_data

    update_log_display("Accept connection")
    btnStart.config(text="Server is running", state=tk.DISABLED)
    btnGameLaoding.config(state=tk.NORMAL)
    while True:
        client, addr = the_server.accept()
        data = client.recv(4096).decode()
        if f"ACCEPT_CON" in data and len(clients_names) == 0:
            clients.append(client)
            # handshaking for the game members
            clients[-1].send("ACCEPT_CON_CC".encode())
            _client = [data, addr]
            clients_names.append(_client)
            update_log_display(f">>> Client is connected <{_client}>")
            # use a thread so as not to clog the gui thread
            _thread.start_new_thread(send_receive_client_message, (client,))
            btnStart.config(text="Get Game results", state=tk.NORMAL)
            break


# Receives a file from client
def send_receive_client_message(client_connected):
    global file_results
    file_data = {"name": None, "size": 0, "start": False, "end": False}
    file = file_bites = None
    while True:
        try:
            data = client_connected.recv(4096)
        except ConnectionResetError:
            break

        print("Status: " + str(file_data))
        if file_data["start"] and file is not None and file_bites is not None:
            file_bites += data
            if file_bites[-5:] == b"<END>":
                file_data["start"] = False
                file_data["end"] = True
                file.write(file_bites)
                file.close()
                update_log_display("FILE COPYING COMPLETE")
                client_connected.send("FILE_COPY_COMPLETE".encode())
        elif "FILE_NAME" in data.decode():
            data = data.decode()
            print(str(data))
            file_data["name"] = str(data).split('::')[-1]
            update_log_display(f"File name: {file_data['name']}")
            client_connected.send("GIVE_FILE_SIZE".encode())
        elif "FILE_SIZE" in data.decode():
            data = data.decode()
            print(str(data))
            file_data["size"] = str(data).split('::')[-1]
            file_data["start"] = True
            update_log_display(f"File size: {file_data['size']}")
            try:
                # file = open(file_data["name"], "wb")
                file = open(file_results, "wb")
                file_bites = b""
                client_connected.send("FILE_SIZE_CC".encode())
            except PermissionError as e:
                print(str(e))
                update_log_display(str(e))
                file_data = {"name": None, "size": 0, "start": False, "end": False}
                file = file_bites = None
                tk.messagebox.showerror(title="File Getting Aborted",
                                        message=f"{e}\nThe file {str(e).split()[-1]} should be closed to be updated")
        elif "FILE_COPY_COMPLETE_CC" in data.decode() and file_data["end"]:
            data = data.decode()
            print(str(data))
            # file_results = file_data["name"]
            file_data = {"name": None, "size": 0, "start": False, "end": False}
            file = file_bites = None
            update_log_display("Ready to get a new file")
            tk.messagebox.showinfo(
                title="RESULT FILE",
                message=f"Excel file '{file_results}' with Result updated")
            # lblFile.config(text=f"File: {file_results}")
        elif not data:
            print("SERVER STOPPED")
            update_log_display("SERVER STOPPED")
            break
        else:
            update_log_display(f"Unknown what to do with the message:\n{data} ")
            pass

    client_idx = get_client_index(clients, client_connected)
    client_connected.close()
    if client_idx >= 0:
        clients.pop(client_idx)
        update_log_display(f"Lost connection with the Client: {clients_names.pop(client_idx)}")
        if len(clients) == 0:
            btnStart.config(text="Re-Start Server", state=tk.NORMAL)
            print("Server can be restarted!!!")
            update_log_display("Server can be restarted!!!")
        else:
            btnStart.config(text="Server is running", state=tk.DISABLED)
            print("Server is running!!!")
            update_log_display("Server is running!!!")
        update_log_display(f"Connected Client is: {len(clients)}")


def message_get_excel_file(client=None):
    """
    Send to Client/s Request to get Excel file with results
    :param client: specific client
    (it by default send the request to all clients are connected to the server)
    """
    if client is not None:
        cc = [client]
    else:
        cc = clients.copy()

    if len(clients) > 0:
        for c in cc:
            c.send(f"GIVE_EXCEL_FILE".encode())
    else:
        update_log_display(f"Request to get file with result impossible since connected Client is: {len(clients)} ")


def get_slide_for_places(pages):
    places = {"1 МЕСТО": None, "ПОБЕДИТЕЛИ!": None, "2 МЕСТО": None, "3 МЕСТО": None,
              "ЗОЛОТАЯ СЕРЕДИНА": None, "ЗАЩИТА ТЫЛА": None, "МЫ СТАРАЛИСЬ!": None}
    for slide in pages:
        # print("Slide index #{}".format(slide.SlideIndex))
        if pages.Count/2 > slide.SlideIndex:
            # prevent scanning 50% slides that is not places
            continue
        for s in slide.Shapes:
            # if "Прямоугольник" in s.Name or "Rectangle" in s.Name:
            #     print("\tShape name: " + s.Name)
            #     print(str(s.TextFrame.TextRange))
            if "Заголовок" in s.Name or "Title" in s.Name:
                # my_str = list_to_str(str(s.TextFrame.TextRange).split())
                print("\tShape name: " + s.Name)
                print(f"<{str(s.TextFrame.TextRange)}>")
                # for p, n in places.items():
                #     if str(s.TextFrame.TextRange) == p and n is None:
                #         print("\tShape name: " + s.Name)
                #         print(str(s.TextFrame.TextRange))
                #         places[p] = slide
                for p, n in places.items():
                    if str(s.TextFrame.TextRange) == p and n is None:
                        print("\tShape name: " + s.Name)
                        print(str(s.TextFrame.TextRange))
                        places[p] = 1
            elif "TextBox" in s.Name:
                # my_str = list_to_str(str(s.TextFrame.TextRange).split())
                print("\tShape name: " + s.Name)
                print(str(s.TextFrame.TextRange))
                for p, n in places.items():
                    if n == 1:
                        print("\tShape name: " + s.Name)
                        print(str(s.TextFrame.TextRange))
                        places[p] = s.TextFrame.TextRange
        for p, n in places.items():
            if n == 1:
                places[p] = None
    return places


def load_places():
    global game_data, game_score, file_results

    if game_data is None:
        file = filedialog.askopenfile(mode='r', filetypes=[('Power Point', '*.pptx')])
        filepath = os.path.abspath(file.name)
        game_data = file
        update_log_display("Game selected:\n" + str(filepath))
        app = win32com.client.Dispatch("PowerPoint.Application")
        # objCOM = app.Presentations.Open(FileName="C:\\Users\\timeo\\Downloads\\GameQ&A#4.pptx", WithWindow=1)
        obj_com = app.Presentations.Open(FileName=filepath, WithWindow=1)
        time.sleep(15)
        update_log_display(f"There are {obj_com.Slides.Count} slides")
        # time.sleep(2)
        # ppt_app = win32com.client.GetActiveObject("PowerPoint.Application")
        # pages = ppt_app.Presentations(1)
        # update_log_display(pages.Name)
        # update_log_display(f"There are {obj_com.Slides.Count} slides")
        game_score = get_slide_for_places(obj_com.Slides)
        if (len(game_score) - sum(1 for v in game_score.values() if v is None)) != 0:  # "None" not in str(game_score):
            btnGameLaoding.config(text="Update places")
        else:
            update_log_display("Not found all the places on the slides in the Game")
            update_log_display(f"{str(game_score)}")
            game_data = None
            game_score = {}
    # elif game_data is not None:  # and len(clients) != 0:
    else:
        if os.path.exists(file_results):
            update_log_display("PLACES IN THE POWER POINT FILE STARTING")
            set_places()
            # for p, n in game_score.items():
            #     print(p)
            #     try:
            #         print(f"<{str(n)}>")
            #         n.Text = "GGGGGG"
            #         print(f"<{str(n)}>")
            #     except Exception as e:
            #         print(f"{e}")
            #         tk.messagebox.showinfo(
            #             title="IMPORTANT",
            #             message=f"PowerPoint Game file is closed!!!\n"
            #                     f"{game_data.name}\n"
            #                     f"YOU MAST OPEN the PowerPoint GAME File again")
            #         close_server(False)
            #         break
        else:
            tk.messagebox.showinfo(title="IMPORTANT",
                                   message=f"File with results didn't find: {file_results}")
        # if os.path.exists(game_data.name):
        #     pass
        #
        # try:
        #     # Call the rename function for the first time
        #     os.rename(filename, newname)
        #     # Call the rename function for the second time
        #     os.rename(filename, newname)
        #     # Raise error if the file has opened
        # except OSError:
        #     print("File is still opened.")


        # game_score = get_places_slide_data(game_data.Slides)
        # for p, n in game_score.items():
        #     print(p)
        #     print(f"<{str(n)}>")

        # for p, n in game_score.items():
        #     for s in n.Shapes:
        #         if "Заголовок" in s.Name or "Title" in s.Name:
        #             print(p)
        #             print("\tShape name: " + s.Name)
        #             print(f"<{str(s.TextFrame.TextRange)}>")
        #         elif "TextBox" in s.Name:
        #             print(p)
        #             print("\tShape name: " + s.Name)
        #             print(str(s.TextFrame.TextRange))
    # else:
    #     update_log_display(f"IMPOSSIBLE UPDATING PLACES YET")
    #     tk.messagebox.showinfo(title="IMPORTANT",
    #                             message="")

def set_places():
    global game_data, game_score, file_results

    if (len(game_score) - sum(1 for v in game_score.values() if v is None)) != 0:  # len(game_score) != 0 and "None" not in str(game_score):
        workbook = openpyxl.load_workbook(file_results)
        sheet = workbook[workbook.sheetnames[0]]
        if sheet.cell(row=1, column=1).value == '#' and sheet.cell(row=1, column=2).value == 'КОМАНДЫ':  # Название
            team_count = 0
            for i in range(2, 50):
                if sheet.cell(row=i, column=2).value is not None:
                    team_count += 1
                else:
                    break
            try:
                messagebox_txt = ""
                if team_count >= 3 and game_score["1 МЕСТО"] is not None:
                    game_score["1 МЕСТО"].Text = sheet.cell(row=2, column=2).value
                    update_log_display(f"1 МЕСТО: {sheet.cell(row=2, column=2).value}")
                    messagebox_txt += f"1 МЕСТО: {game_score['1 МЕСТО'].Text}\n"
                elif team_count >= 3 and game_score["ПОБЕДИТЕЛИ!"] is not None:
                    game_score["ПОБЕДИТЕЛИ!"].Text = sheet.cell(row=2, column=2).value
                    update_log_display(f"ПОБЕДИТЕЛИ!: {sheet.cell(row=2, column=2).value}")
                    messagebox_txt += f"ПОБЕДИТЕЛИ!: {game_score['ПОБЕДИТЕЛИ!'].Text}\n"
                else:
                    update_log_display(f"1 МЕСТО или ПОБЕДИТЕЛИ! не найдены в PowerPoint файле")
                    messagebox_txt += f"1 МЕСТО или ПОБЕДИТЕЛИ! не добавлен  в PowerPoint файле\n"

                if team_count >= 3 and game_score["2 МЕСТО"] is not None:
                    game_score["2 МЕСТО"].Text = sheet.cell(row=3, column=2).value
                    update_log_display(f"2 МЕСТО: {sheet.cell(row=3, column=2).value}")
                    messagebox_txt += f"2 МЕСТО: {game_score['2 МЕСТО'].Text}\n"
                else:
                    update_log_display(f"2 МЕСТО не найден в PowerPoint файле")
                    messagebox_txt += f"2 МЕСТО не добавлен в PowerPoint файле\n"

                if team_count >= 3 and game_score["3 МЕСТО"] is not None:
                    game_score["3 МЕСТО"].Text = sheet.cell(row=4, column=2).value
                    update_log_display(f"3 МЕСТО: {sheet.cell(row=4, column=2).value}")
                    messagebox_txt += f"3 МЕСТО: {game_score['3 МЕСТО'].Text}\n"
                else:
                    update_log_display(f"3 МЕСТО не найден в PowerPoint файле")
                    messagebox_txt += f"3 МЕСТО не добавлен в PowerPoint файле\n"

                if game_score["ПОБЕДИТЕЛИ!"] is None and team_count >= 7:
                    if team_count % 2 != 0:
                        game_score["ЗОЛОТАЯ СЕРЕДИНА"].Text = sheet.cell(row=team_count // 2 + 2, column=2).value
                        update_log_display(f"ЗОЛОТАЯ СЕРЕДИНА: {sheet.cell(row=team_count // 2 + 2, column=2).value}")
                        messagebox_txt += f"ЗОЛОТАЯ СЕРЕДИНА: {game_score['ЗОЛОТАЯ СЕРЕДИНА'].Text}\n"
                    else:
                        game_score["ЗОЛОТАЯ СЕРЕДИНА"].Text = sheet.cell(row=team_count // 2 + 1, column=2).value
                        update_log_display(f"ЗОЛОТАЯ СЕРЕДИНА: {sheet.cell(row=team_count // 2 + 1, column=2).value}")
                        messagebox_txt += f"ЗОЛОТАЯ СЕРЕДИНА: {game_score['ЗОЛОТАЯ СЕРЕДИНА'].Text}\n"
                else:
                    update_log_display(f"ЗОЛОТАЯ СЕРЕДИНА не номенируеться в этой игре или если"
                                       f" играет меньше 7и команд (играет {team_count} команд)")
                    messagebox_txt += f"ЗОЛОТАЯ СЕРЕДИНА не номенируеться (играет {team_count} команд)\n"

                if team_count >= 5 and game_score["ЗАЩИТА ТЫЛА"] is not None:
                    game_score["ЗАЩИТА ТЫЛА"].Text = sheet.cell(row=team_count, column=2).value
                    update_log_display(f"ЗАЩИТА ТЫЛА: {sheet.cell(row=team_count, column=2).value}")
                    messagebox_txt += f"ЗАЩИТА ТЫЛА: {game_score['ЗАЩИТА ТЫЛА'].Text}"
                elif team_count >= 4 and game_score["МЫ СТАРАЛИСЬ!"] is not None:
                    game_score["МЫ СТАРАЛИСЬ!"].Text = sheet.cell(row=team_count+1, column=2).value
                    update_log_display(f"МЫ СТАРАЛИСЬ!: {sheet.cell(row=team_count+1, column=2).value}")
                    messagebox_txt += f"МЫ СТАРАЛИСЬ!: {game_score['МЫ СТАРАЛИСЬ!'].Text}"
                else:
                    update_log_display(f"ЗАЩИТА ТЫЛА или МЫ СТАРАЛИСЬ! не найдены в PowerPoint файле")
                    messagebox_txt += f"ЗАЩИТА ТЫЛА или МЫ СТАРАЛИСЬ! не добавлен в PowerPoint файле\n"

                tk.messagebox.showinfo(title="PLACES SETUP NOTIFICATION",
                                       message=messagebox_txt)
            except Exception as e:
                print(f"{e}")
                tk.messagebox.showinfo(
                    title="IMPORTANT",
                    message=f"PowerPoint Game file is closed!!!\n"
                            f"{game_data.name}\n"
                            f"YOU MAST OPEN the PowerPoint GAME File again")
                close_server(False)


def close_server(with_verification: bool = True):
    answer = 'yes'
    if with_verification:
        answer = tk.messagebox.askquestion(f"Confirm", f"Are you sure you want close the connection")
    if answer == 'yes':
        for c in clients:
            c.send("ServerClose".encode())
        _thread.exit()


def get_client_index(client_list, curr_client):
    """Return ID or -1 if the current client didn't find"""
    idx = 0
    for conn in client_list:
        if conn == curr_client:
            return idx
        idx = idx + 1
    return -1


def update_log_display(text):
    tkDisplay.config(state=tk.NORMAL)
    if text != '' or text is not None:
        tkDisplay.insert(tk.END, text + "\n")
    tkDisplay.config(state=tk.DISABLED)


window.mainloop()
