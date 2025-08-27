import tkinter as tk
from tkinter import messagebox
import socket
import win32com.client
import openpyxl

import _thread
import json
import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

server = None
file_results = "results.xlsx" # result table file name

# 0/RED=Wrong, 1/GREEN=Good and 2/Yellow=None
# bg_result = ["#F61F03", "#2FFFC1", "#F6FCC7"]
# HOST_ADDR = socket.gethostbyname(socket.gethostname())  # 0.0.0.0
HOST_ADDR = socket.gethostname()
print(f"Host IP: {socket.gethostname()}")
print(f"Host IP: {socket.gethostbyname(socket.gethostname())}")
HOST_PORT = 8080
print(f"Port used: {HOST_PORT}")

client_name = " "
clients = []
clients_names = []
player_data = []


def open_socket():
    """ Open server socket """
    global server, HOST_ADDR, HOST_PORT  # , teams  # code is fine without this

    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    print(socket.AF_INET)
    print(socket.SOCK_STREAM)
    server.bind((HOST_ADDR, HOST_PORT))
    server.listen(1)  # server is listening for client connection


def list_to_str(data: list):
    """
     Convert a list to a string with specific formatting.
    :param data: (list) The list to convert.
    :returns: (str) A string representation of the list with elements separated by spaces,
                or with special characters '/' and '=' treated as separators.
    """
    my_str = ''
    if len(data) > 0:
        for i in data:
            if my_str == '' or i == '/' or i == '=' or i == '+' or my_str[-1] == '/':
                my_str += i
            else:
                my_str += f" {i}"
    return my_str


window = tk.Tk()
window.title("Game Server")

# open server socket
open_socket()
text = f"HOST: {socket.gethostname()}/{socket.gethostbyname(socket.gethostname())}\nPort: {HOST_PORT}"
btnFrame = tk.Frame(window)
btnStart = tk.Button(btnFrame, text="Start Server", command=lambda: start_server())
btnStart.pack(side='left')
btnSetPlaces = tk.Button(btnFrame, text="Set Places", font=("Arial Bold", 10), bg="orange", fg="red",
                         command=lambda: load_places(), state='disabled')
btnSetPlaces.pack(side='left')
btnStop = tk.Button(btnFrame, text="Close Server", command=lambda: close_server(), state='normal')
btnStop.pack(side='left')
btnFrame.pack(side='top', pady=(5, 0))

# Middle frame consisting of two labels for displaying the host and port info
middleFrame = tk.Frame(window)
lblHost = tk.Label(middleFrame, text=f"Address: {HOST_ADDR}")
lblHost.pack(side='left')
lblPort = tk.Label(middleFrame, text=f"Port: {HOST_PORT}")
lblPort.pack(side='left')
lblFile = tk.Label(middleFrame, text=f"File: {file_results}")
lblFile.pack(side='left')
middleFrame.pack(side='top', pady=(5, 0))

# Game table
gameFrame = tk.Frame(window)
gameFrame.pack(side='top', pady=(5, 0))

# The client frame shows the client area
clientFrame = tk.Frame(window)
tk.Label(clientFrame, text="********** LOG **********").pack()
scrollBar = tk.Scrollbar(clientFrame)
scrollBar.pack(side='right', fill='y')
tkDisplay = tk.Text(clientFrame, height=10, width=100)
tkDisplay.pack(side='left', fill='y', padx=(5, 0))
scrollBar.config(command=tkDisplay.yview)
tkDisplay.config(
    yscrollcommand=scrollBar.set,
    background="#F4F6F7",
    highlightbackground="grey",
    state="disabled",
)
clientFrame.pack(side='bottom', pady=(5, 10))
if text != '' or text is not None:
    tkDisplay.config(state='normal')
    tkDisplay.insert(tk.END, text + "\n")
tkDisplay.config(state='disabled')


def start_server():
    """
     Accept client connections if no clients are connected,
      otherwise send a request to get the Excel file.
    """
    if len(clients_names) == 0:
        # accept connection a first client required
        _thread.start_new_thread(accept_client_connection, (server,))
    else:
        message_get_excel_file()
        pass


def accept_client_connection(the_server):
    """
    Accept a client connection and handle the handshaking process.
    :param the_server: The server socket to accept connections from a client.
    """
    update_log_display("Accept connection")
    btnStart.config(text="Server is running", state='disabled')
    btnSetPlaces.config(state='normal')
    while True:
        client, addr = the_server.accept()
        data = client.recv(4096).decode()
        if f"ACCEPT_CON" in data and len(clients_names) == 0:
            clients.append(client)
            # handshaking for the game members
            clients[-1].send("ACCEPT_CON_CC".encode())
            _client = [data, addr]
            clients_names.append(_client)
            update_log_display(f">>> Client is connected <{_client}>")
            # use a thread so as not to clog the gui thread
            _thread.start_new_thread(send_receive_client_message, (client,))
            btnStart.config(text="Get Game results", state='normal')
            break


# Receives a file from client
def send_receive_client_message(client_connected):
    """
    Handle messages from a connected client.
    This function listens for messages from the client, processes file transfers
     and updates the log display accordingly.
    :param client_connected: The connected client socket.
    """
    global file_results
    file_data = {"name": None, "size": 0, "start": False, "end": False}
    file = file_bites = None
    while True:
        try:
            data = client_connected.recv(4096)
        except ConnectionResetError:
            break

        print("Status: " + str(file_data))
        if file_data["start"] and file is not None and file_bites is not None:
            file_bites += data
            if file_bites[-5:] == b"<END>":
                file_data["start"] = False
                file_data["end"] = True
                file.write(file_bites)
                file.close()
                update_log_display("FILE COPYING COMPLETE")
                client_connected.send("FILE_COPY_COMPLETE".encode())
        elif "FILE_NAME" in data.decode():
            data = data.decode()
            print(str(data))
            file_data["name"] = str(data).split('::')[-1]
            update_log_display(f"File name: {file_data['name']}")
            client_connected.send("GIVE_FILE_SIZE".encode())
        elif "FILE_SIZE" in data.decode():
            data = data.decode()
            print(str(data))
            file_data["size"] = str(data).split('::')[-1]
            file_data["start"] = True
            update_log_display(f"File size: {file_data['size']}")
            try:
                # file = open(file_data["name"], "wb")
                file = open(file_results, "wb")
                file_bites = b""
                client_connected.send("FILE_SIZE_CC".encode())
            except PermissionError as e:
                print(str(e))
                update_log_display(str(e))
                file_data = {"name": None, "size": 0, "start": False, "end": False}
                file = file_bites = None
                tk.messagebox.showerror(title="File Getting Aborted",
                                        message=f"{e}\nThe file {str(e).split()[-1]} should be closed to be updated")
        elif "FILE_COPY_COMPLETE_CC" in data.decode() and file_data["end"]:
            data = data.decode()
            print(str(data))
            # file_results = file_data["name"]
            file_data = {"name": None, "size": 0, "start": False, "end": False}
            file = file_bites = None
            update_log_display("Ready to get a new file")
            tk.messagebox.showinfo(
                title="RESULT FILE",
                message=f"Excel file '{file_results}' with Result updated")
            # lblFile.config(text=f"File: {file_results}")
        elif not data:
            print("SERVER STOPPED")
            update_log_display("SERVER STOPPED")
            break
        else:
            update_log_display(f"Unknown what to do with the message:\n{data} ")
            pass

    client_idx = get_client_index(clients, client_connected)
    client_connected.close()
    if client_idx >= 0:
        clients.pop(client_idx)
        update_log_display(f"Lost connection with the Client: {clients_names.pop(client_idx)}")
        if len(clients) == 0:
            btnStart.config(text="Re-Start Server", state='normal')
            print("Server can be restarted!!!")
            update_log_display("Server can be restarted!!!")
        else:
            btnStart.config(text="Server is running", state='disabled')
            print("Server is running!!!")
            update_log_display("Server is running!!!")
        update_log_display(f"Connected Client is: {len(clients)}")


def message_get_excel_file(client=None):
    """
    Send to Client/s Request to get Excel file with results
    :param client: specific client
    (it by default send the request to all clients are connected to the server)
    """
    if client is not None:
        cc = [client]
    else:
        cc = clients.copy()

    if len(clients) > 0:
        for c in cc:
            c.send(f"GIVE_EXCEL_FILE".encode())
    else:
        update_log_display(f"Request to get file with result impossible since connected Client is: {len(clients)} ")


def load_places():
    """ Load places from the Excel file and set them in the opened PowerPoint file."""
    global file_results
    if os.path.exists(file_results):
        update_log_display("\n!!! UPDATING PLACES IN THE POWER POINT FILE !!!\n")
        res = set_places_with_teams_in_opened_pptx(team_list=get_teams_ordered_from_excel_file(file_results))
        if res is None:
            tk.messagebox.showerror(title="PLACES SETUP ERROR",
                                    message="PLaces were not set !!!!!")
        elif "ERROR" in res.upper():
            tk.messagebox.showerror(title="PLACES SETUP ERROR",
                                    message=res)
        else:
            tk.messagebox.showinfo(title="PLACES SETUP NOTIFICATION",
                                   message=res)
    else:
        tk.messagebox.showerror(title="IMPORTANT",
                               message=f"The Result File didn't find:\n {file_results}")

def is_item_in_list_or_nesty_lists(key_item:str, lists_with_items:list):
    """
    Check if a key_item is in a list or in any nested lists within lists_with_items.
    :param key_item: (str) The item to search for.
    :param lists_with_items: (list) A list that may contain other lists or items.
    :returns: (bool) True if key_item is found, False otherwise.
    """
    for item in lists_with_items:
        if isinstance(item, list):
            if is_item_in_list_or_nesty_lists(key_item, item):
                return True
        else:
            if key_item == item:
                return True
    return False


def get_item_index(key_item:str, lists_with_items:list):
    """
    Get the index of a key_item in a list or in any nested lists within lists_with_items.
    This function searches for the key_item in the provided list and its nested lists.
    :param key_item: (str) The item to search for.
    :param lists_with_items: (list) A list that may contain other lists or items.
    :returns: If the key_item is found, it returns the index of the first occurrence.
                If the key_item is not found, it returns -1.
    """
    index_of_key_item = -1  # Default value if key_item is not found
    try:
        index_of_key_item = lists_with_items.index(key_item)
    except ValueError:
        for item_index in range(len(lists_with_items)):
            if isinstance(lists_with_items[item_index], list):
                if is_item_in_list_or_nesty_lists(key_item, lists_with_items[item_index]):
                    return item_index
    return index_of_key_item

def get_teams_ordered_from_excel_file(file_results="results.xlsx"):
    """
    Load teams from an Excel file and return them as a list.
    This function reads the first sheet of the specified Excel file and looks for the secretive words:
      1. Looks for '# in' row 1 and column 1
      2. Looks for 'КОМАНДЫ' in row 1 and column 2
      3. Extracts team names from each raw in column 2 onwards until an empty cell is found
    :param file_results: (str) Path to the Excel file containing team names.
    :returns: (list) a list of team names is ordered by results.
    """
    workbook = openpyxl.load_workbook(file_results)
    sheet_name = 'AllResultsOnTable'  # default name for the first sheet required
    update_log_display(f"Looks for sheet: '{sheet_name}' in the file:\n" + str(file_results))
    print(f"Looks for sheet: '{sheet_name}' in the file:" + str(file_results))
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]

    # sheet = workbook[workbook.sheetnames[0]]
    team_list = []
    update_log_display("Looks for 'КОМАНДЫ' column")
    print("Looks for 'КОМАНДЫ' column")
    if sheet.cell(row=1, column=1).value == '#' and sheet.cell(row=1, column=2).value == 'КОМАНДЫ':  # Название
        for i in range(2, 50):
            if sheet.cell(row=i, column=2).value is not None:
                team_list.append(sheet.cell(row=i, column=2).value)
            else:
                break
    update_log_display("Team list: " + str(team_list))
    print("Team list: " + str(team_list))
    return team_list

def set_places_with_teams_in_opened_pptx(team_list:list, config_path="places.json"):
    """
    Set places with teams in an opened PowerPoint file based on a configuration file.
    :param config_path: (str) Path to the configuration file containing definition for
      ["PLACES", "EVEN", "ODD", "LAST"].
    :param team_list: (list) a list of team names is ordered by results.
    :returns: (str or None) Error message if any error occurs, otherwise None.
    """
    ERR0 = "Error: No config file found or it is not a valid JSON."
    ERR1 = "Error: No PLACES key found in the config file or not enough keys."
    ERR2 = "Error: No opened PowerPoint file found."
    ERR3 = "Error: Expected exactly one opened PowerPoint file, found {}."
    ERR4 = "Error: Not enough teams provided. At least 3 teams are required."

    # Load config
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)
    config_keys = config.get("PLACES", [])

    if not config or not isinstance(config, dict):
        update_log_display(ERR0)
        print(ERR0)
        return ERR0
    if len(config_keys) < 3:
        update_log_display(ERR1)
        print(ERR0)
        return ERR1

    # Set up teams into middle and last places:
    # -> middle for 7 or more teams
    #  -> last for 4 or 5 teams
    _middle = _last = None
    if len(team_list) >= 7:
        if config.get("EVEN") and len(team_list) % 2 == 0:
            _middle = team_list[int(len(team_list)/2-1)]
        elif config.get("ODD") and len(team_list) % 2 != 0:
            _middle = team_list[int(len(team_list)/2)]

    if config.get("LAST") and len(team_list) >= 4:
        _last = team_list[int(len(team_list)-1)]
    elif not config.get("LAST") and len(team_list) >= 5:
        _last = team_list[int(len(team_list)-2)]
    update_log_display(f"Configuration loaded properly from: {config_path}")
    print(f"Configuration loaded properly from: {config_path}")

    # Connect to opened PowerPoint
    try:
        app = win32com.client.GetActiveObject("PowerPoint.Application")
        update_log_display("Connected to PowerPoint")
        print("Connected to PowerPoint")
    except Exception as e:
        update_log_display(f"Error connecting to PowerPoint: {e}")
        update_log_display(ERR2)
        print(f"Error connecting to PowerPoint: {e}")
        return ERR2

    if app.Presentations.Count != 1:
        # raise RuntimeError(f"Expected exactly one opened PowerPoint file, found {app.Presentations.Count}.")
        update_log_display(ERR3.format(app.Presentations.Count))
        print(ERR3.format(app.Presentations.Count))
        return ERR3.format(app.Presentations.Count)

    if len(team_list)< 3:
        update_log_display(ERR4)
        print(ERR4)
        return ERR4

    presentation = app.ActivePresentation
    slide_count = presentation.Slides.Count
    update_log_display(f"Slides in file: {slide_count}")
    print(f"Slides in file: {slide_count}")
    if slide_count < 1:
        update_log_display("No slides found in the presentation.")
        print("No slides found in the presentation.")
        return "Error: No slides found in the presentation."

    messagebox_txt = None
    for i in range(slide_count // 2 + 1, slide_count + 1):
        slide = presentation.Slides(i)
        update_log_display("Scanning Slide #: " + str(i))
        print("Scanning Slide #: " + str(i))
        title_found = -1
        title_text = ''
        for shape in slide.Shapes:
            if "Заголовок" in shape.Name or "Title" in shape.Name:
                title_text = str(shape.TextFrame.TextRange)
                title_found = get_item_index(title_text, config_keys)
                if title_found >= 0:  # text in str(word_keys):
                    update_log_display(f"Found key '{title_text}' on slide {i}")
                    # print(str(shape.TextFrame.TextRange))
                else:
                    title_found = -1
            elif title_found >= 0 and "TextBox" in shape.Name:
                """ If the title was found, look for the next TextBox to set the team name """
                # print("\tShape name: " + shape.Name)
                # print(str(shape.TextFrame.TextRange))
                if title_found < 3:
                    shape.TextFrame.TextRange.Text = team_list[title_found]
                elif title_found == 3 and _middle is not None:
                    shape.TextFrame.TextRange.Text = _middle
                elif title_found > 3 and _last is not None:
                    shape.TextFrame.TextRange.Text = _last
                else:
                    continue
                res = str(shape.TextFrame.TextRange.Text)
                update_log_display(
                    f"Wining Place '{title_text}' by team: '{res}'")
                print(f"Wining Place '{title_text}' by team: '{res}'")
                if messagebox_txt is None:
                    messagebox_txt = f"'{title_text}' ==> '{res}'\n"
                else:
                    messagebox_txt += f"'{title_text}' ==> '{res}'\n"
    return messagebox_txt


def close_server(with_verification: bool = True):
    """ Close the server and all client connections """
    answer = 'yes'
    if with_verification:
        answer = tk.messagebox.askquestion(f"Confirm", f"Are you sure you want close the connection")
    if answer == 'yes':
        for c in clients:
            c.send("ServerClose".encode())
        _thread.exit()


def get_client_index(client_list, curr_client):
    """Return ID or -1 if the current client didn't find"""
    idx = 0
    for conn in client_list:
        if conn == curr_client:
            return idx
        idx = idx + 1
    return -1


def update_log_display(text):
    """ Update the log display with new text """
    tkDisplay.config(state='normal')
    if text != '' or text is not None:
        tkDisplay.insert(tk.END, text + "\n")
    tkDisplay.config(state='disabled')


window.mainloop()
