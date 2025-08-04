import json
import customtkinter as ctk
from tkinter import filedialog, messagebox
import openpyxl
from telethon.sync import TelegramClient
from telethon.tl.types import InputPeerUser
import re
import asyncio
import threading
import os  # Already imported, but explicitly for os.environ

# Set appearance mode and default color theme for customtkinter
ctk.set_appearance_mode("System")  # Modes: "System" (default), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (default), "green", "dark-blue"


class TelegramNotifierApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Telegram Notifier App")
        self.geometry("1000x700")

        # Configure grid layout (2 columns for panels)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- Input Panel ---
        self.input_panel = ctk.CTkFrame(self)
        self.input_panel.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        # Input Panel grid configuration:
        self.input_panel.grid_columnconfigure(0, weight=1)
        self.input_panel.grid_columnconfigure(1, weight=1)

        # Configure rows for Input Panel elements
        self.input_panel.grid_rowconfigure(0, weight=0)  # Title
        self.input_panel.grid_rowconfigure(1, weight=0)  # Telegram API ID label
        self.input_panel.grid_rowconfigure(2, weight=0)  # Telegram API ID entry
        self.input_panel.grid_rowconfigure(3, weight=0)  # Telegram API Hash label
        self.input_panel.grid_rowconfigure(4, weight=0)  # Telegram API Hash entry
        self.input_panel.grid_rowconfigure(5, weight=0)  # Excel File label and button
        self.input_panel.grid_rowconfigure(6, weight=0)  # Excel file path display
        self.input_panel.grid_rowconfigure(7, weight=0)  # Common Message label
        self.input_panel.grid_rowconfigure(8, weight=1)  # Message input (expands vertically)
        self.input_panel.grid_rowconfigure(9, weight=0)  # New row for the two preset message buttons
        self.input_panel.grid_rowconfigure(10, weight=0)  # Send button (moved to new row)

        ctk.CTkLabel(self.input_panel, text="Input Panel", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0,
                                                                                                          column=0,
                                                                                                          pady=(0, 20))

        # Telegram Credentials
        ctk.CTkLabel(self.input_panel, text="Telegram API ID:").grid(row=1, column=0, columnspan=2, sticky="w", padx=10,
                                                                     pady=(0, 5))
        self.api_id_entry = ctk.CTkEntry(self.input_panel, placeholder_text="Enter your API ID")
        self.api_id_entry.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 5))
        self.api_id_entry.bind("<KeyRelease>", self._check_send_button_state)

        ctk.CTkLabel(self.input_panel, text="Telegram API Hash:").grid(row=3, column=0, columnspan=2, sticky="w",
                                                                       padx=10, pady=(10, 5))
        self.api_hash_entry = ctk.CTkEntry(self.input_panel, placeholder_text="Enter your API Hash")
        self.api_hash_entry.grid(row=4, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 5))
        self.api_hash_entry.bind("<KeyRelease>", self._check_send_button_state)

        # Excel File Selection
        ctk.CTkLabel(self.input_panel, text="Excel File:").grid(row=5, column=0, sticky="w", padx=10, pady=(20, 5))
        self.excel_file_path_label = ctk.CTkLabel(self.input_panel, text="No file selected")
        self.excel_file_path_label.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 10))
        self.select_excel_button = ctk.CTkButton(self.input_panel, text="Select Excel File",
                                                 command=self._select_excel_file)
        self.select_excel_button.grid(row=5, column=1, sticky="e", padx=10, pady=(20, 10))

        # Free Text Message Input
        ctk.CTkLabel(self.input_panel, text="Common Message:").grid(row=7, column=0, columnspan=2, sticky="nw", padx=10,
                                                                    pady=(20, 5))
        self.message_textbox = ctk.CTkTextbox(self.input_panel, height=150)
        self.message_textbox.grid(row=8, column=0, columnspan=2, sticky="nsew", padx=10, pady=(5, 10))
        # Default message template with placeholders
        self.message_textbox.insert("0.0", "Здравствуйте, '{team_name}' !!!")
        self.message_textbox.bind("<KeyRelease>", self._check_send_button_state)  # Bind for validation

        self._load_message_templates()
        self.message_names = list(self.message_templates.keys())
        self.message_combobox = ctk.CTkComboBox(
            self.input_panel,
            values=self.message_names,
            command=self._on_message_selected,
            state="readonly"  # Prevents text entry
        )
        self.message_combobox.grid(row=9, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 10))
        self.message_combobox.set(self.message_names[0])  # Set default

        # Send Notification Button
        self.send_button = ctk.CTkButton(self.input_panel, text="Send Notifications", command=self._send_notifications,
                                         state="disabled")
        self.send_button.grid(row=10, column=0, columnspan=2, pady=20)

        # --- Control Panel ---
        self.control_panel = ctk.CTkFrame(self)
        self.control_panel.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        self.control_panel.grid_columnconfigure(0, weight=1)
        self.control_panel.grid_rowconfigure(0, weight=0)
        self.control_panel.grid_rowconfigure(1, weight=1)
        self.control_panel.grid_rowconfigure(2, weight=0)
        self.control_panel.grid_rowconfigure(3, weight=0)
        self.control_panel.grid_rowconfigure(4, weight=0)

        ctk.CTkLabel(self.control_panel,
                     text="Control Panel",
                     font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(0,20))

        # User List Display
        self.user_scroll_frame = ctk.CTkScrollableFrame(self.control_panel,
                                                        label_text="Parsed Users (Select to Send)")
        self.user_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.user_scroll_frame.grid_columnconfigure(0, weight=1)

        self.user_checkboxes = {}

        # Select/Deselect all buttons
        self.select_all_button = ctk.CTkButton(self.control_panel, text="Select All",
                                               command=lambda: self._toggle_all_users(True))
        self.select_all_button.grid(row=2, column=0, sticky="w", padx=10, pady=(5, 0))
        self.deselect_all_button = ctk.CTkButton(self.control_panel, text="Deselect All",
                                                 command=lambda: self._toggle_all_users(False))
        self.deselect_all_button.grid(row=2, column=0, sticky="e", padx=10, pady=(5, 0))

        # Status/Log Area
        ctk.CTkLabel(self.control_panel, text="Status/Log:").grid(row=3, column=0, sticky="w", padx=10, pady=(10, 5))
        self.log_textbox = ctk.CTkTextbox(self.control_panel, height=150, state="disabled")
        self.log_textbox.grid(row=4, column=0, sticky="nsew", padx=10, pady=(5, 10))

        # Internal state variables
        self.excel_file_path = None
        self.parsed_data = []  # List of dicts: {"team_name": ..., "result": ..., "phone_number": ...}
        self.telegram_client = None
        self.telegram_session_file = 'telegram_notifier_session.session'  # Telethon session file

        # Call initial state check to disable send button
        self._check_send_button_state()

        # --- NEW: Try to get API credentials from environment variables ---
        telegram_api_id = os.environ.get('TELEGRAM_API_ID')
        telegram_api_hash = os.environ.get('TELEGRAM_API_HASH')

        if telegram_api_id:
            self.api_id_entry.insert(0, telegram_api_id)
            self._log_message("API ID pre-filled from TELEGRAM_API_ID environment variable.")
        if telegram_api_hash:
            self.api_hash_entry.insert(0, telegram_api_hash)
            self._log_message("API Hash pre-filled from TELEGRAM_API_HASH environment variable.")
        # --- END NEW --

    def _load_message_templates(self):
        default_messages = {
            "Уведомление о результате": "Здравствуйте, '{team_name}', ваш результат:\n{result}",
            "Подтверждение присутствия": "Подтвердите свое присутствие на игре в команде '{team_name}'"
        }
        try:
            with open("messages.json", "r", encoding="utf-8") as f:
                extra_messages = json.load(f)
                default_messages.update(extra_messages)
        except Exception:
            pass  # Ignore if file not found or invalid

        self.message_templates = default_messages

    def _on_message_selected(self, selected_name):
        message = self.message_templates.get(selected_name, "")
        self.message_textbox.delete("1.0", "end")
        self.message_textbox.insert("1.0", message)
        self._check_send_button_state()

    def _log_message(self, message):
        """Appends a message to the log textbox and scrolls to the end."""
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", message + "\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")

    def _check_send_button_state(self, event=None):
        """Checks if all conditions are met to enable the 'Send Notifications' button."""
        api_id_valid = False
        try:
            if self.api_id_entry.get():
                int(self.api_id_entry.get())  # Check if it's an integer
                api_id_valid = True
        except ValueError:
            api_id_valid = False

        api_hash_valid = bool(self.api_hash_entry.get())
        # Button is enabled if file is selected, parsed data exists, and credentials/message are present.
        file_and_data_valid = self.excel_file_path is not None and len(self.parsed_data) > 0
        message_present = bool(self.message_textbox.get("1.0", "end-1c").strip())

        if api_id_valid and api_hash_valid and file_and_data_valid and message_present:
            self.send_button.configure(state="normal")
        else:
            self.send_button.configure(state="disabled")

    def _select_excel_file(self):
        """Opens a file dialog to select an Excel file and triggers parsing."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_path = file_path
            self.excel_file_path_label.configure(text=f"Selected: {os.path.basename(self.excel_file_path)}")
            self._log_message(f"Selected Excel file: {self.excel_file_path}")
            self._parse_excel_file()
        else:
            self.excel_file_path = None
            self.excel_file_path_label.configure(text="No file selected")
            self._log_message("Excel file selection cancelled.")
            self.parsed_data = []  # Clear data if selection cancelled or failed
            self._update_user_list_display()  # Clear display
            self._check_send_button_state()

    def _parse_excel_file(self):
        """Parses the selected Excel file according to specified rules."""
        self.parsed_data = []
        errors = []
        warnings = []  # New list to store informational messages
        team_results = {}
        team_numbers = {}

        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)

            # Check for at least two sheets
            if len(workbook.sheetnames) < 2:
                errors.append("Excel file must have at least two sheets for data.")
                raise ValueError("Not enough sheets")

            # --- Parse Sheet 1: Team Names + Results ---
            sheet1_name = workbook.sheetnames[0]
            sheet1 = workbook[sheet1_name]
            headers1 = [str(cell.value).upper().strip() if cell.value is not None else "" for cell in sheet1[1]]
            try:
                team_name_col_idx1 = headers1.index("КОМАНДЫ")  # Find the starting index for "КОМАНДЫ"
            except ValueError:
                errors.append(f"Sheet 1 ('{sheet1_name}'): Missing 'КОМАНДЫ' in header in the first row.")
                team_name_col_idx1 = -1

            if team_name_col_idx1 != -1:
                for row_idx in range(2, sheet1.max_row + 1):
                    result_parts = []
                    team = sheet1.cell(row=row_idx, column=team_name_col_idx1 + 1).value
                    if team is None:
                        continue

                    col_idx = team_name_col_idx1 + 2  # Start from the next column after "КОМАНДЫ"
                    while col_idx <= sheet1.max_column:
                        header = sheet1.cell(row=1, column=col_idx).value
                        if header is None or str(header).strip() == "":
                            break  # Stop at the first empty header cell
                        value = sheet1.cell(row=row_idx, column=col_idx).value
                        result_parts.append(f"{str(header).strip()}: {value}")
                        col_idx += 1

                    result = "\n".join(result_parts)
                    team_results[str(team).strip()] = result if result else "N\\A"

            # --- Parse Sheet 2: Team Names + Telegram Phone Numbers ---
            sheet2_name = workbook.sheetnames[1]
            sheet2 = workbook[sheet2_name]
            if str(sheet2['A1'].value).upper().strip() != "КОМАНДЫ":
                errors.append(f"Sheet 2 ('{sheet2_name}'): Cell A1 must be 'КОМАНДЫ'.")
            else:
                headers2 = [cell.value for cell in sheet2[1]]
                try:
                    team_name_col_idx2 = headers2.index("КОМАНДЫ")  # Case-sensitive
                    phone_number_col_idx = headers2.index("Phone Number")  # Case-sensitive
                except ValueError:
                    errors.append(
                        f"Sheet 2 ('{sheet2_name}'): Missing 'КОМАНДЫ' or 'Phone Number' header in the first row.")
                    team_name_col_idx2 = -1  # Indicate failure to find
                    phone_number_col_idx = -1

                if team_name_col_idx2 != -1 and phone_number_col_idx != -1:
                    for row_idx in range(2, sheet2.max_row + 1):
                        team = sheet2.cell(row=row_idx, column=team_name_col_idx2 + 1).value
                        phone = sheet2.cell(row=row_idx, column=phone_number_col_idx + 1).value
                        if team and phone:
                            phone_str = str(phone).strip()
                            # Basic phone number validation: starts with +, then 10-15 digits
                            if re.fullmatch(r'^\+\d{10,15}$', phone_str):
                                team_numbers[str(team).strip()] = phone_str
                            else:
                                warnings.append(
                                    f"WARNING: Invalid phone number format for team '{team}': '{phone_str}'. Must start with '+' and be 10-15 digits long (e.g., +12345678900). This team will not be included.")

            # --- Combine and Validate Data with Info/Warnings ---
            # Teams that exist in both sheets (candidates for sending)
            common_teams = set(team_results.keys()).intersection(set(team_numbers.keys()))

            for team_name in sorted(list(common_teams)):  # Sort for consistent display
                # We already know they exist in both, so get will return a value
                result = team_results[team_name]
                phone_number = team_numbers[team_name]
                self.parsed_data.append({
                    "team_name": team_name,
                    "result": result,
                    "phone_number": phone_number
                })

            # Report teams found only in Sheet 1 (results but no phone)
            sheet1_only_teams = set(team_results.keys()) - common_teams
            for team_name in sorted(list(sheet1_only_teams)):
                warnings.append(
                    f"INFO: Team '{team_name}' found in Sheet 1 (Results) but missing in Sheet 2 (Phone Numbers)."
                    f" Will not send message to this team.")

            # Report teams found only in Sheet 2 (phone but no result)
            sheet2_only_teams = set(team_numbers.keys()) - common_teams
            for team_name in sorted(list(sheet2_only_teams)):
                warnings.append(
                    f"INFO: Team '{team_name}' found in Sheet 2 (Phone Numbers) but missing in Sheet 1 (Results)."
                    f" Will not send message to this team.")

            if not self.parsed_data and not errors and not warnings:  # Only if literally nothing was found
                errors.append("No team names found in either sheet, or no complete valid data for any team.")

        except FileNotFoundError:
            errors.append(f"Excel file not found at: {self.excel_file_path}")
        except Exception as e:
            errors.append(f"An unexpected error occurred during Excel parsing: {type(e).__name__}: {e}")

        if errors:
            messagebox.showerror("Excel Parsing Error", "\n".join(errors))
            self._log_message("Excel parsing failed with critical errors:\n" + "\n".join(errors))
            self.parsed_data = []  # Ensure no partial data is used if critical errors occurred
        else:
            if warnings:
                self._log_message("Excel parsing completed with warnings:")
                for warning in warnings:
                    self._log_message(warning)
            self._log_message(
                f"Successfully parsed Excel file. Found {len(self.parsed_data)} valid entries for sending messages.")
            if not self.parsed_data and not warnings:  # If no data and no warnings, it means parsing was technically fine but found nothing
                messagebox.showwarning("No Data Found",
                                       "Excel file parsed, but no complete team data (name, result, phone) was found for sending messages.")

        self._update_user_list_display()
        self._check_send_button_state()

    def _update_user_list_display(self):
        """Clears and repopulates the user list display in the Control Panel."""
        # Clear existing checkboxes
        for widget in self.user_scroll_frame.winfo_children():
            widget.destroy()
        self.user_checkboxes.clear()

        if not self.parsed_data:
            ctk.CTkLabel(self.user_scroll_frame, text="No users parsed yet.").grid(row=0, column=0, padx=5, pady=5)
            return

        # Populate with new data
        for i, user_data in enumerate(self.parsed_data):
            var = ctk.BooleanVar(value=True)  # All selected by default
            checkbox = ctk.CTkCheckBox(self.user_scroll_frame,
                                       text=f"{user_data['team_name']} ({user_data['phone_number']})",
                                       variable=var)
            checkbox.grid(row=i, column=0, sticky="w", padx=5, pady=2)
            self.user_checkboxes[user_data['team_name']] = {"checkbox": checkbox, "var": var, "data": user_data}

    def _toggle_all_users(self, select_all: bool):
        """Selects or deselects all user checkboxes."""
        for user_info in self.user_checkboxes.values():
            user_info["var"].set(select_all)

    async def _run_telegram_send_async(self):
        """Asynchronous function to handle Telegram connection and message sending."""
        api_id_str = self.api_id_entry.get()
        api_hash = self.api_hash_entry.get()
        message_template = self.message_textbox.get("1.0", "end-1c")

        try:
            api_id = int(api_id_str)
        except ValueError:
            self._log_message("ERROR: Telegram API ID must be a valid integer.")
            messagebox.showerror("Input Error", "Telegram API ID must be a valid integer.")
            return

        if not api_hash:
            self._log_message("ERROR: Telegram API Hash cannot be empty.")
            messagebox.showerror("Input Error", "Telegram API Hash cannot be empty.")
            return

        selected_users = [
            user_info["data"] for user_info in self.user_checkboxes.values()
            if user_info["var"].get()
        ]

        if not selected_users:
            self._log_message("WARNING: No users selected to send notifications to.")
            messagebox.showwarning("No Selection", "No users selected to send notifications to.")
            return

        self._log_message("Connecting to Telegram...")
        try:
            self.telegram_client = TelegramClient(self.telegram_session_file, api_id, api_hash)
            await self.telegram_client.connect()

            if not await self.telegram_client.is_user_authorized():
                self._log_message(
                    "AUTHORIZATION REQUIRED: Please check your console/terminal"
                    " for a Telegram login prompt (phone number and code).")
                # This will prompt for phone number and code if not authorized.
                # It might block the console until authorized.
                await self.telegram_client.start()
                self._log_message("Telegram client authorized successfully.")

            self._log_message(f"Attempting to send messages to {len(selected_users)} selected users...")
            for user_data in selected_users:
                team_name = user_data["team_name"]
                result = user_data["result"]
                phone_number = user_data["phone_number"]

                # Format the message using parsed data
                formatted_message = message_template.format(team_name=team_name, result=result)

                try:
                    # Resolve user by phone number. This requires the user to be discoverable
                    # or in the contacts of the Telegram account logged in.
                    entity = await self.telegram_client.get_input_entity(phone_number)
                    await self.telegram_client.send_message(entity, formatted_message)
                    self._log_message(f"SUCCESS: Sent message to {team_name} ({phone_number}).")
                except ValueError as ve:
                    self._log_message(
                        f"ERROR: Could not resolve user entity for {team_name} ({phone_number}). ({ve})."
                        f" Ensure phone number is correct and user is discoverable.")
                except Exception as e:
                    self._log_message(f"ERROR: Failed to send message to {team_name} ({phone_number}): {e}")
            self._log_message("Finished attempting to send all messages.")

        except Exception as e:
            self._log_message(f"CRITICAL ERROR: Failed to connect or send messages via Telegram: {e}")
            messagebox.showerror(
                "Telegram Error", f"Failed to connect or send messages via Telegram: {e}")
        finally:
            if self.telegram_client and self.telegram_client.is_connected():
                await self.telegram_client.disconnect()
                self._log_message("Disconnected from Telegram.")

    def _send_notifications(self):
        """Starts the asynchronous Telegram sending process in a separate thread to avoid blocking the GUI."""

        # This function is called by the GUI button.
        # We need to run the async Telethon code in a separate thread to not freeze the CustomTkinter GUI.

        def run_async_in_thread(coro):
            """Helper function to run an async coroutine in a new event loop on a separate thread."""
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                loop.run_until_complete(coro)
            finally:
                loop.close()

        # Create and start a new thread for the Telegram operations
        thread = threading.Thread(target=run_async_in_thread, args=(self._run_telegram_send_async(),))
        thread.daemon = True  # Allow the thread to exit when the main app exits
        thread.start()
        self._log_message(
            "Initiated sending process in background."
            " Check log for real-time updates and console for authorization.")


if __name__ == "__main__":
    app = TelegramNotifierApp()
    app.mainloop()
